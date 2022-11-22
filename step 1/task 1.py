import csv
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, numbers


class DictByCity:
    def __init__(self):
        self.data_dict = {}
        self.total_count = 0

    def add_data(self, city, value):
        if city not in self.data_dict.keys():
            self.data_dict[city] = [value, 1]
        else:
            self.data_dict[city][0] += value
            self.data_dict[city][1] += 1
        self.total_count += 1

    def get_by_salary(self):
        sub_dict = {}
        sort_dic = dict(sorted(self.data_dict.items(), key=lambda item: item[1][0] // item[1][1], reverse=True))
        for key, value in sort_dic.items():
            ratio = value[1] / self.total_count
            if ratio >= 0.01:
                sub_dict[key] = int(value[0] // value[1])
            if len(sub_dict) == 10:
                break

        return sub_dict

    def get_by_count(self):
        sub_dict = {}
        sort_dic = dict(sorted(self.data_dict.items(), key=lambda item: item[1][1], reverse=True))
        for key, value in sort_dic.items():
            ratio = value[1] / self.total_count
            if ratio >= 0.01:
                sub_dict[key] = round(ratio, 4)
            if len(sub_dict) == 10:
                break
        return sub_dict


class DictByYear:
    def __init__(self):
        self.data_dict = {}

    def add_data(self, year, value, count):
        if year not in self.data_dict.keys():
            self.data_dict[year] = [value, count]
        else:
            self.data_dict[year][0] += value
            self.data_dict[year][1] += count

    def get_by_salary(self):
        sub_dict = {}
        for key, value in self.data_dict.items():
            sub_dict[key] = int(value[0] // value[1]) if value[1] != 0 else 0
        return sub_dict

    def get_by_count(self):
        sub_dict = {}
        for key, value in self.data_dict.items():
            sub_dict[key] = value[1]
        return sub_dict


class DataSet:
    def __init__(self, file_name):
        self.file_name = file_name
        self.vacancies_objects = DataSet._set_vacancies(*DataSet._csv_reader(file_name))

        self.salary_count_by_year = DictByYear()
        self.job_salary_count_by_year = DictByYear()
        self.salary_count_by_city = DictByCity()

    @staticmethod
    def _csv_reader(file_name):
        file_csv = csv.reader(open(file_name, encoding='utf_8_sig'))
        list_data = [x for x in file_csv if x.count('') == 0]

        return list_data[0], list_data[1:]

    @staticmethod
    def _set_vacancies(headers, vacancies):
        list_vacancies = []
        for vac in vacancies:
            dic = {}
            for i, item in enumerate(headers):
                dic[item] = vac[i]
            list_vacancies.append(
                Vacancy(dic['name'], dic['salary_from'], dic['salary_to'], dic['salary_currency'], dic['area_name'],
                        dic['published_at']))

        return list_vacancies

    def collect_statistic(self, profession):
        for vac in self.vacancies_objects:
            self.salary_count_by_year.add_data(vac.published_at.year, vac.average_salary, 1)
            if profession in vac.name:
                self.job_salary_count_by_year.add_data(vac.published_at.year, vac.average_salary, 1)
            else:
                self.job_salary_count_by_year.add_data(vac.published_at.year, 0, 0)
            self.salary_count_by_city.add_data(vac.area_name, vac.average_salary)

    def get_statistic(self):
        return self.salary_count_by_year.get_by_salary(), \
               self.salary_count_by_year.get_by_count(), \
               self.job_salary_count_by_year.get_by_salary(), \
               self.job_salary_count_by_year.get_by_count(), \
               self.salary_count_by_city.get_by_salary(), \
               self.salary_count_by_city.get_by_count()

    def print_statistic(self):
        print(f"Динамика уровня зарплат по годам: {self.salary_count_by_year.get_by_salary()}")
        print(f"Динамика количества вакансий по годам: {self.salary_count_by_year.get_by_count()}")
        print(
            f"Динамика уровня зарплат по годам для выбранной профессии: {self.job_salary_count_by_year.get_by_salary()}")
        print(
            f"Динамика количества вакансий по годам для выбранной профессии: {self.job_salary_count_by_year.get_by_count()}")
        print(f"Уровень зарплат по городам (в порядке убывания): {self.salary_count_by_city.get_by_salary()}")
        print(f"Доля вакансий по городам (в порядке убывания): {self.salary_count_by_city.get_by_count()}")


class Vacancy:
    _currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, name, salary_from, salary_to, salary_currency, area_name, published_at):
        self.name = name
        self.average_salary = (float(salary_from) + float(salary_to)) / 2 * Vacancy._currency_to_rub[
            salary_currency]
        self.area_name = area_name
        self.published_at = datetime.datetime.strptime(published_at, "%Y-%m-%dT%H:%M:%S%z")


class Report:
    def __init__(self, salary_by_year, count_by_year, job_salary_by_year, job_count_by_year,
                 salary_by_city, count_by_city, prof):
        self.profession = prof
        self.salary_by_year = salary_by_year
        self.count_by_year = count_by_year
        self.profession_salary_by_year = job_salary_by_year
        self.profession_count_by_year = job_count_by_year
        self.salary_by_city = salary_by_city
        self.count_by_city = count_by_city

    @staticmethod
    def _as_text(value):
        if value is None:
            return ""
        return str(value)

    def generate_excel(self):
        wb = Workbook()
        thin = Side(border_style="thin", color="000000")
        self._create_first_tab(wb, thin)
        self._create_second_tab(wb, thin)

        wb.save("report.xlsx")

    def _create_first_tab(self, workbook, pen):
        tab1 = workbook.active
        tab1.title = "Статистика по годам"
        heads = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.profession}", "Количество вакансий",
                 f"Количество вакансий - {self.profession}"]
        for i, head in enumerate(heads):
            tab1.cell(row=1, column=(1 + i), value=head).font = Font(bold=True)
        for year, value in self.salary_by_year.items():
            tab1.append([year, value, self.profession_salary_by_year[year], self.count_by_year[year],
                         self.profession_count_by_year[year]])
        Report._set_length_and_border(tab1, pen)

    def _create_second_tab(self, workbook, pen):
        tab2 = workbook.create_sheet("Статистика по городам")
        heads = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        for i, head in enumerate(heads):
            tab2.cell(row=1, column=1 + i, value=head).font = Font(bold=True)
        for i, (city, value) in enumerate(self.salary_by_city.items()):
            city_count, value_count = list(self.count_by_city.items())[i]
            tab2.append([city, value, "", city_count, value_count])
            tab2.cell(row=2+i, column=5).number_format = numbers.BUILTIN_FORMATS[10]
        Report._set_length_and_border(tab2, pen, True)

    @staticmethod
    def _set_length_and_border(tab, pen, is_city=False):
        for i, column in enumerate(tab.columns):
            length = max(len(Report._as_text(cell.value)) for cell in column)
            tab.column_dimensions[column[0].column_letter].width = length + 2
            for cell in column:
                if is_city and i == 2:
                    break
                cell.border = Border(left=pen, right=pen, top=pen, bottom=pen)


file_name = input('Введите название файла: ')
profession = input('Введите название профессии: ')
data = DataSet(file_name)
data.collect_statistic(profession)
data.print_statistic()

rep = Report(*data.get_statistic(), profession)
rep.generate_excel()
